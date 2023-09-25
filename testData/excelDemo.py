import openpyxl
filepath = "D:/selenium/PythonDemo.xlsx"
book = openpyxl.load_workbook(filepath)

sheet = book.active
Dict = {}
cell = sheet.cell(row=1, column=2)
print(cell.value)
name_input = sheet.cell(row=3, column=2)
print(sheet['A5'].value)
name_input.value = "Bhargavi Golla"
print(name_input.value)
# book.save(filepath)

print(sheet.max_row)
print(sheet.max_column)

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(2,sheet.max_column+1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i,column=j).value

print(Dict)
