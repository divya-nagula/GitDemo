import openpyxl


class HomePageData:
    test_HomePage_data = [{"name": "Divya Nagula", "Email": "divyanagula@gmail.com", "gender": "Female"},
                          {"name": "Bhargavi Golla", "Email": "bhargavi@gmail.com", "gender": "Female"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        filepath = "D:/selenium/PythonDemo.xlsx"
        book = openpyxl.load_workbook(filepath)
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[Dict]
